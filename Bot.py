# from ogame import OGame
# from ogame.constants import Ships, Speed, Facilities, Buildings, Research, Defense, Missions
from pyogame.ogame import OGame
from ogame.constants import Ships, Speed, Facilities, Buildings, Research, Defense, Missions
import Utils
import openpyxl
from time import sleep
import random
from tinydb import TinyDB, Query, where
from datetime import datetime, date
from dateutil import parser


class PlanetInfo:
    def __init__(self, planet_id, ogame):
        self.id = planet_id
        self.resources = ogame.get_resources(planet_id)
        self.resources_buildings = ogame.get_resources_buildings(planet_id)
        self.planet_overview = ogame.get_overview(planet_id)
        self.facilities = ogame.get_facilities(planet_id)
        self.level = self.resources_buildings['metal_mine'] + \
                     self.resources_buildings['crystal_mine'] + \
                     self.resources_buildings['solar_plant'] + \
                     self.resources_buildings['deuterium_synthesizer']
        self.infos = ogame.get_planet_infos(planet_id)   # TO DELETE
        self.ships = ogame.get_ships(planet_id)
        self.defense = ogame.get_defense(planet_id)


class Bot:
    def __init__(self, login, password, server, uni):
        self.collaborate_minimum_level = 75
        self.login = login
        self.password = password
        self.server = server
        self.uni = uni
        self.ogame = OGame(self.uni, self.login, self.password, self.server)

        self.mother_id = self.ogame.get_planet_by_name('Planeta matka')
        self.planets = self.ogame.get_planet_ids()
        self.researches = self.ogame.get_research()
        self.weakest_planet_id = 0
        self.planet_infos = {}
        for planet_id in self.planets:
            print('Loading planet {}'.format(planet_id))
            self.planet_infos[planet_id] = PlanetInfo(planet_id, self.ogame)

        colony_doc = openpyxl.load_workbook('colony_planet_build.xlsx')
        colony_sheet = colony_doc.active
        self.colony_cells = colony_sheet['A1': 'C139']

        mother_doc = openpyxl.load_workbook('first_planet_build.xlsx')
        mother_sheet = mother_doc.active
        self.mother_cells = mother_sheet['A1': 'C185']

        shared_doc = openpyxl.load_workbook('shared_build.xlsx')
        shared_sheet = shared_doc.active
        self.shared_cells = shared_sheet['A1': 'C200']

        self.res_req_db = TinyDB('resources_requests.json')

    def start_def(self):
        for planet_info in self.planet_infos.values():
            planet_id = planet_info.id
            print('p Planet currently working: ' + planet_info.infos['planet_name'])
            start_time = datetime.now()
            try:
                if planet_id == self.mother_id:
                    self.develop_mother_planet(planet_info, self.mother_cells)
                else:
                    self.develop_colony_planet(planet_info, self.colony_cells)
            except Exception as e:
                print('-----======------- Error happend in start_def -----======-------')
                print(e)

            time_elapsed = datetime.now() - start_time
            sleep_time = random.uniform(0.5, 1.3)
            print('p Planet handling time elapsed {}, random sleep time: {}[ms]'.format(time_elapsed, sleep_time))
            sleep(sleep_time)

    def start_fleet(self):
        pass

    def start_eco(self):
        planets = self.planets
        researches = self.ogame.get_research()
        colony_slots = researches['astrophysics'] / 2

        if colony_slots > len(planets) - 1:
            self.colonize_planet_eco(self.planet_infos[self.mother_id])

        for planet_info in self.planet_infos.values():
            planet_id = planet_info.id
            print('Planet currently working: ' + planet_info.infos['planet_name'])
            try:
                if planet_id == self.mother_id:
                    self.develop_mother_planet(planet_info, self.mother_cells)
                else:
                    self.develop_colony_planet(planet_info, self.colony_cells)
            except Exception as e:
                print('-----======------- Error happend in  start_eco -----======-------')
                print(e)
            sleep_time = random.uniform(0.5, 1.4)
            sleep(sleep_time)
            print('p Planet {} handling finished, random sleep time: {}[ms]'.
                  format(planet_info.infos['planet_name'], sleep_time))

    @staticmethod
    def check_anti_ballistic_missiles_on_main_planet(ogame):
        mother_id = ogame.get_planet_by_name('Planeta matka')

        defense = ogame.get_defense(mother_id)
        defend_rockets = defense['anti_ballistic_missiles']
        if defend_rockets < 50:
            print('##### ALERT!! REBUILDING ROCKET DEFENSE: ' + defend_rockets)
            ogame.build(mother_id, (Defense['AntiBallisticMissiles'], 70))

    def develop_colony_planet(self, planet_info, colony_cells):
        self.check_colony_def_ships(planet_info)
        next_build = self.get_next_build(planet_info, colony_cells)
        self.build_on_planet(planet_info, next_build)
        if planet_info.level > self.collaborate_minimum_level:
            if len(planet_info.planet_overview['buildings']) == 0:
                self.request_resources_for_next_build(planet_info, next_build)
            self.collect_resources(self.mother_id, planet_info)

    def develop_mother_planet(self, mother_info, mother_cells):
        self.check_mother_defense_and_ships(mother_info)
        self.send_expeditions(mother_info)
        next_build = self.get_next_build(mother_info, mother_cells)
        self.build_on_planet(mother_info, next_build)
        if mother_info.level > self.collaborate_minimum_level:
            self.send_resources_from_mother_if_possible(mother_info, self.res_req_db)

    def colonize_planet_eco(self, planet_info):
        if planet_info.ships['colony_ship'] < 1:
            if len(planet_info.planet_overview['shipyard']) > 0:
                return
            self.ogame.build(self.mother_id, (Ships['ColonyShip'], 1))
            return
        mother_coordinates = planet_info.infos['coordinate']
        start_galaxy = mother_coordinates['galaxy']
        # start_system = mother_coordinates['system']
        start_system = random.randint(100, 400)
        resources = {'metal': 0, 'crystal': 0, 'deuterium': 0}
        slots = [11, 10, 9, 8, 7, 6, 5]
        current_system = start_system

        while True:
            for slot in slots:
                destination = {'galaxy': start_galaxy, 'system': current_system, 'position': slot}
                self.ogame.send_fleet(self.mother_id, [(Ships['ColonyShip'], 1)], Speed['100%'],
                                      destination, Missions['Colonize'], resources)

                if self.ogame.get_ships(self.mother_id)['colony_ship'] < 1:
                    print('skolonizowano: ' + str(destination))
                    return

            system_step = start_system - current_system
            if system_step == 0:
                current_system += 1
            elif system_step < 0:
                current_system = start_system + system_step
            elif system_step > 0:
                current_system = start_system + system_step + 1

            if system_step > 50:
                return

    def build_on_planet(self, planet_info, build_info):
        if build_info is None:
            return
        researches = self.researches
        resources_buildings = planet_info.resources_buildings
        civilian_buildings = planet_info.facilities

        if build_info[0] == '':
            return
        elif build_info[0] in resources_buildings and resources_buildings[build_info[0]] < build_info[1]:
            self.ogame.build(planet_info.id, Buildings[build_info[2]])
        elif build_info[0] in civilian_buildings and civilian_buildings[build_info[0]] < build_info[1]:
            self.ogame.build(planet_info.id, Facilities[build_info[2]])
        elif build_info[0] in researches and researches[build_info[0]] < build_info[1]:
            self.ogame.build(planet_info.id, Research[build_info[2]])

    def get_next_build(self, planet_info, build_cells):
        researches = self.researches
        resources_buildings = planet_info.resources_buildings
        civilian_buildings = planet_info.facilities

        for c1, c2, c3 in build_cells:
            if c1.value == '':
                return
            elif c1.value in resources_buildings and resources_buildings[c1.value] < c2.value:
                return [c1.value, c2.value, c3.value]
            elif c1.value in civilian_buildings and civilian_buildings[c1.value] < c2.value:
                return [c1.value, c2.value, c3.value]
            elif c1.value in researches and researches[c1.value] < c2.value:
                return [c1.value, c2.value, c3.value]

    def check_minimum_large_cargos(self, planet_info):
        if len(planet_info.planet_overview['shipyard']) > 0:
            return
        ships = planet_info.ships
        dt_count = 0

        if planet_info.level > 70:
            dt_count = 5
        if planet_info.level > 80:
            dt_count = 15

        dt = int(ships['large_cargo'])
        if dt < dt_count:
            self.ogame.build(planet_info.id, (Ships['LargeCargo'], dt_count - dt))

    def check_colony_def_ships(self, planet_info):
        if len(planet_info.planet_overview['shipyard']) > 0:
            return

        defense = planet_info.defense
        ships = planet_info.ships
        facilities = planet_info.facilities
        planet_id = planet_info.id

        silos_level = int(facilities['missile_silo'])
        ldl = int(defense['light_laser'])
        gauses = int(defense['gauss_cannon'])
        solars = int(ships['solar_satellite'])
        probes = int(ships['espionage_probe'])
        dt = int(ships['large_cargo'])
        shipyard = int(facilities['shipyard'])

        ldl_count = 0
        anti_rockets_count = 0
        gauss_count = 0
        solars_count = 0
        probes_count = 0
        dt_count = 0

        if planet_info.level > self.collaborate_minimum_level:
            dt_count = 15 - dt
            probes_count = 1 - probes
            ldl_count = 200 - ldl
        if shipyard >= 8 and planet_info.level > 80:
            ldl_count = 500 - ldl
            gauss_count = 20 - gauses
            solars_count = 100 - solars
            dt_count = 20 - dt
        if silos_level >= 2:
            anti_rockets_count = silos_level * 10

        if anti_rockets_count > 0:
            self.ogame.build(planet_id, (Defense['AntiBallisticMissiles'], anti_rockets_count))
        if gauss_count > 0:
            self.ogame.build(planet_id, (Defense['GaussCannon'], gauss_count))
        if ldl_count > 0:
            self.ogame.build(planet_id, (Defense['LightLaser'], ldl_count))
        if solars_count > 0:
            self.ogame.build(planet_id, (Ships['SolarSatellite'], solars_count))
        if probes_count > 0:
            self.ogame.build(planet_id, (Ships['EspionageProbe'], probes_count))
        if dt_count > 0 and probes_count > 0:
            self.ogame.build(planet_id, (Ships['LargeCargo'], dt_count))

    def check_mother_defense_and_ships(self, planet_info):
        if len(planet_info.planet_overview['shipyard']) > 0:
            return

        defense = planet_info.defense
        ships = planet_info.ships
        resources_buildings = planet_info.resources_buildings
        facilities = planet_info.facilities
        planet_id = planet_info.id

        silos_level = int(facilities['missile_silo'])
        ldl = int(defense['light_laser'])
        gauses = int(defense['gauss_cannon'])
        anti_rockets = int(defense['anti_ballistic_missiles'])
        solars = int(ships['solar_satellite'])
        probes = int(ships['espionage_probe'])
        dt = int(ships['large_cargo'])
        metal_mine = int(resources_buildings['metal_mine'])
        crystal_mine = int(resources_buildings['crystal_mine'])
        shipyard = int(facilities['shipyard'])

        ldl_count = 0
        anti_rockets_count = 0
        gauss_count = 0
        solars_count = 0
        probes_count = 0
        dt_count = 0

        if planet_info.level > self.collaborate_minimum_level:
            dt_count = 15 - dt
            probes_count = 1 - probes

        if crystal_mine > 15:
            ldl_count = 113 - ldl
            probes_count = 1 - probes
        if shipyard >= 8 and planet_info.level > 80:
            ldl_count = 1200 - ldl
            gauss_count = 30
            dt_count = 45 - dt
        if shipyard >= 8 and planet_info.level > 90:
            ldl_count = 1000 - ldl
            gauss_count = 50 - gauses
            solars_count = 150 - solars
            dt_count = 80 - dt
        if silos_level >= 2:
            anti_rockets_count = silos_level * 10

        if anti_rockets_count > 0:
            self.ogame.build(planet_id, (Defense['AntiBallisticMissiles'], anti_rockets_count))
        if gauss_count > 0:
            self.ogame.build(planet_id, (Defense['GaussCannon'], gauss_count))
        if ldl_count > 0:
            self.ogame.build(planet_id, (Defense['LightLaser'], ldl_count))
        if solars_count > 0:
            self.ogame.build(planet_id, (Ships['SolarSatellite'], solars_count))
        if probes_count > 0:
            self.ogame.build(planet_id, (Ships['EspionageProbe'], probes_count))
        if dt_count > 0:
            self.ogame.build(planet_id, (Ships['LargeCargo'], dt_count))

    def build_on_planet_from_parsed_xlsx(self, planet_id, cells, planetInfo):
        researches = self.researches
        # print('DEBUG: c1, c2, c3 in cells:')
        for c1, c2, c3 in cells:
            if c1.value == '':
                return
            elif c1.value in planetInfo.resources_buildings and planetInfo.resources_buildings[c1.value] < c2.value:
                if len(planetInfo.planet_overview['buildings']) > 0:
                    continue
                print('DEBUG: building build: ' + c1.value + ' on level: ' + str(c2.value))
                self.ogame.build(planet_id, Buildings[c3.value])
                return
            elif c1.value in planetInfo.facilities and planetInfo.facilities[c1.value] < c2.value:
                if len(planetInfo.planet_overview['buildings']) > 0:
                    continue
                print('DEBUG: building facilities: ' + c1.value + ' on level: ' + str(c2.value))
                self.ogame.build(planet_id, Facilities[c3.value])
                return
            elif c1.value in researches and researches[c1.value] < c2.value:
                if len(planetInfo.planet_overview['research']) > 0:
                    continue
                print('DEBUG: building technology: ' + c1.value + ' on level: ' + str(c2.value))
                self.ogame.build(planet_id, Research[c3.value])
                return

    def send_resources_from_mother_if_possible(self, mother_info, res_req_db):
        requests = res_req_db.all()
        if requests is None:
            return
        lowest_request = 99999999999
        working_request = None
        for request in requests:
            requesting_id = request['requesting_id']
            was_sent = request['sent']
            if request['login'] != self.login or request['uni'] != self.uni or \
                            request['requesting']['galaxy'] != mother_info.infos['coordinate']['galaxy']:
                continue
            if was_sent == 'true' or len(self.ogame.get_overview(requesting_id)['buildings']) > 0:
                continue
            req_metal = request['metal']
            req_crystal = request['crystal']
            req_deuterium = request['deuterium']
            req_total_cost = req_metal + req_crystal + 1.4*req_deuterium

            # find lowest request
            if req_total_cost < lowest_request:
                lowest_request = req_total_cost
                working_request = request

        if working_request is None:
            return

        req_metal = working_request['metal']
        req_crystal = working_request['crystal']
        req_deuterium = working_request['deuterium']
        req_building = working_request['building']
        req_level = working_request['level']
        total_required_resources = req_metal + req_crystal + req_deuterium

        metal = mother_info.resources['metal']
        crystal = mother_info.resources['crystal']
        deuterium = mother_info.resources['deuterium']

        if metal > req_metal and crystal > req_crystal and deuterium > req_deuterium:
            dt_count = int(total_required_resources / 25000) + 1
            ships = [(Ships['LargeCargo'], dt_count)]
            if mother_info.ships['large_cargo'] < dt_count:
                return
            speed = Speed['100%']
            where_to = {'galaxy': working_request['requesting']['galaxy'],
                     'system': working_request['requesting']['system'],
                     'position': working_request['requesting']['position']}
            mission = Missions['Transport']
            request_resources = {'metal': req_metal, 'crystal': req_crystal, 'deuterium': req_deuterium}
            self.ogame.send_fleet(mother_info.id, ships, speed, where_to, mission, request_resources)
            res_req_db.remove(where('requesting') == working_request['requesting'] and
                              where('metal') == req_metal and
                              where('crystal') == req_crystal and
                              where('uni') == self.uni)
            res_req_db.insert({'requesting': working_request['requesting'], 'metal': req_metal, 'crystal': req_crystal,
                               'deuterium': req_deuterium, 'building': working_request['building'],
                               'level': working_request['level'], 'login': self.login,
                               'uni': self.uni, 'sent': 'true', 'requesting_id': working_request['requesting_id']})
            print('-> sending resources from mother to: ' + str(where_to))

    def collect_resources(self, mother_id, planet_info):
        dt = planet_info.ships['large_cargo']
        capacity = dt * 25000
        resources = planet_info.resources['metal'] + \
                    planet_info.resources['crystal'] + \
                    planet_info.resources['deuterium']
        if resources > capacity * 0.75 and dt > 5:
            ships = [(Ships['LargeCargo'], 9999)]
            speed = Speed['100%']
            where = self.planet_infos[mother_id].infos['coordinate']
            mission = Missions['Transport']
            resources = {'metal': 999999999, 'crystal': 99999999, 'deuterium': 99999999}
            self.ogame.send_fleet(planet_info.id, ships, speed, where, mission, resources)
            print('collect_resources from: ' + planet_info.infos['planet_name'] + ' to mother planet')

    def request_starter_resources(self, planet_info, res_req_db):
        exst_req = res_req_db.search(Query().requesting == planet_info.infos['coordinate'])
        if len(exst_req) > 0:
            return

        res_req_db.insert({'requesting': planet_info.infos['coordinate'], 'metal': 50000, 'crystal': 50000,
                           'deuterium': 20000})

    def send_expedition(self, planet_info):
        mother_coordinates = planet_info.infos['coordinate']
        galaxy = mother_coordinates['galaxy']
        system = mother_coordinates['system']
        system_delta = random.randint(-5, 5)
        dt_count = 0

        if planet_info.level > 70:
            dt_count = 5
        if planet_info.level > 80:
            dt_count = 15
        if planet_info.level > 95:
            dt_count = 100
        if planet_info.level > 100:
            dt_count = 250

        ships = [(Ships['EspionageProbe'], 1), (Ships['LargeCargo'], dt_count)]
        speed = Speed['100%']
        where = {'galaxy': galaxy, 'system': system + system_delta, 'position': 16}
        mission = Missions['Expedition']
        resources = {'metal': 0, 'crystal': 0, 'deuterium': 0}
        self.ogame.send_fleet(planet_info.id, ships, speed, where, mission, resources)
        print('send expedition from mother to: {}'.format(where))

    def send_expeditions(self, planet_info):
        mother_coordinates = planet_info.infos['coordinate']
        galaxy = mother_coordinates['galaxy']
        system = mother_coordinates['system']
        for i in (1, 2, 3):
            system_delta = random.randint(-5, 5)
            dt_count = 0

            if planet_info.level > 70:
                dt_count = 5
            if planet_info.level > 80:
                dt_count = 15
            if planet_info.level > 95:
                dt_count = 100
            if planet_info.level > 100:
                dt_count = 250

            ships = [(Ships['EspionageProbe'], 1), (Ships['LargeCargo'], dt_count)]
            speed = Speed['100%']
            where = {'galaxy': galaxy, 'system': system + system_delta, 'position': 16}
            mission = Missions['Expedition']
            resources = {'metal': 0, 'crystal': 0, 'deuterium': 0}
            self.ogame.send_fleet(planet_info.id, ships, speed, where, mission, resources)
            print('attemp {}: send expedition from mother to: {}'.format(i, where))

    def request_resources_for_next_build_old(self, planet_info, res_req_db, shared_cells):
        exst_req = res_req_db.search(Query().requesting == planet_info.infos['coordinate'])
        if exst_req is None:
            return
        if len(exst_req) > 0:
            return
        # print('DEBUG: for c1, c2, c3 in shared_cells:')
        for c1, c2, c3 in shared_cells:
            if c1.value is None or c1.value == '' or c1.value == 'solar_satellite':
                continue
            # print('DEBUG: Utils.calc_build_cost({}, {}):'.format(c1.value, c2.value))
            cost = Utils.calc_build_cost(c1.value, c2.value)
            building = c1.value
            level = c2.value
            if building in planet_info.facilities and planet_info.facilities[building] >= level:
                continue
            if building in planet_info.resources_buildings and planet_info.resources_buildings[building] >= level:
                continue
            res_req_db.insert({'requesting': planet_info.infos['coordinate'], 'metal': cost[0], 'crystal': cost[1],
                               'deuterium': cost[2], 'building': building, 'level': level, 'login': self.login,
                               'uni': self.uni, 'sent': 'false', 'requesting_id': planet_info.id})

    def request_resources_for_next_build(self, planet_info, next_build):
        if next_build is None:
            return
        exst_req = self.res_req_db.search(Query().requesting == planet_info.infos['coordinate'])
        if exst_req is None:
            return
        if len(exst_req) > 0:
            return
        if next_build[0] is None or next_build[0] == '' or next_build[0] == 'solar_satellite':
            return

        cost = Utils.calc_build_cost(next_build[0], next_build[1])
        building = next_build[0]
        level = next_build[1]

        self.res_req_db.insert({'requesting': planet_info.infos['coordinate'], 'metal': cost[0], 'crystal': cost[1],
                           'deuterium': cost[2], 'building': building, 'level': level, 'login': self.login,
                           'uni': self.uni, 'sent': 'false', 'requesting_id': planet_info.id})

    def build_next_build(self, planet_info, shared_cells):
        pass

    def check_anti_ballistic_missiles(self):
        for planet_id in self.planet_infos:
            name = self.ogame.get_planet_infos(planet_id)['planet_name']
            if 'matka' in name:
                # print('defending: ' + name)
                defense = self.ogame.get_defense(planet_id)
                defend_rockets = defense['anti_ballistic_missiles']
                if defend_rockets < 50:
                    self.ogame.build(planet_id, (Defense['AntiBallisticMissiles'], 70))
